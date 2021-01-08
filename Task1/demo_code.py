# -*- coding: utf-8 -*-
"""
Created on Thu Jan  7 14:45:48 2021

@author: HP
"""

from  tkinter import *
import tkinter as tk 
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

main=Tk()
main.title("details")
main.geometry("800x500")
main.config(highlightbackground="black",highlightthickness=2)

file=pathlib.Path("demo.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="User Name"
    sheet["C1"]="Password"
    sheet["D1"]="Email"
    sheet["E1"]="Gender"
    sheet["F1"]="Subscription"
    
    file.save("demo.xlsx")
    

        
        
def submit():
    y=name.get()
    z=user.get()
    z1=passEntry.get()
    y1=emailentry.get()
    print(y)
    print(z)
    print(z1)
    print(y1)
    
    file=openpyxl.load_workbook("demo.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=z1)
    sheet.cell(column=4,row=sheet.max_row,value=y1)
    
    
    if var1.get()==1:
        gen="male"
        print("Male")
        sheet.cell(column=5,row=sheet.max_row,value="male")
    else:
        print("Female")
        sheet.cell(column=5,row=sheet.max_row,value="female")
         
    
    if var2.get()==1:
       
        print("STD")
        sheet.cell(column=6,row=sheet.max_row,value="STD")
        
        
    if var3.get()==1:
        print("Pre")
        sheet.cell(column=6,row=sheet.max_row,value="Premium") 

    file.save("demo.xlsx")
        


frame1=LabelFrame(main,text='login details').pack(expand='yes',fill='both')
Label(frame1,text="Name").place(x=50,y=30)
Label(frame1,text="username").place(x=50,y=70)
Label(frame1,text="Password").place(x=50,y=110)
Label(main,text="mail id").place(x=50,y=150)

frame2=LabelFrame(main,text='oter details').pack(expand='yes',fill='both')
name=Entry(frame1)
name.place(x=250,y=30)

user=Entry(frame1)
user.place(x=250,y=70)

password=StringVar()
passEntry=Entry(frame1,textvariable=password,show='*')
passEntry.place(x=250,y=110)

emailentry=Entry(frame1)
emailentry.place(x=250,y=150,width=250)

''' var=IntVar()
var.set('0')
chkbutton=Checkbutton(frame1,text='Show',variable=var,onvalue=1,offvalue=0,command=show_password).place(x=350,y=110) '''

label_3=Label(frame2,text="Gender",width=20,font=("bold",10))
label_3.place(x=50,y=300)
var1=IntVar()
Radiobutton(frame2,text="male",padx=5,variable=var1,value=1).place(x=200,y=300)
Radiobutton(frame2,text="Fe-male",padx=20,variable=var1,value=2).place(x=280,y=300)

label_4=Label(frame2,text="subsc",width=20,font=("bold",10))
label_4.place(x=50,y=360)

var2=IntVar()
Checkbutton(frame2,text="std",variable=var2).place(x=200,y=360)
var3=IntVar()
Checkbutton(frame2,text="pre",variable=var3).place(x=300,y=360) 

Button(frame2,text="susb",command=submit).place(x=400,y=400)
main.mainloop()
        

    

         
    
    
  
         
    
    
    

    

